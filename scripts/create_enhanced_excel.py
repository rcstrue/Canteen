#!/usr/bin/env python3
"""
Create enhanced RCS Canteen Stock Excel with all missing data from JSON backup.
Adds: Supplier Master, Ingredient Master, Recipe, Recipe Ingredient, Budget, Expense, Stock Movement, Daily Meal Served, Audit Log, User sheets.
Preserves original sheets: Master, Purchase, Use, Meal, Stock Report.
"""

import json
import sys
import os
from datetime import datetime
from copy import copy

# --- Skill imports ---
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from base import (
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    CHART_COLORS, COLUMN_WIDTHS, FORMATS, ROW_HEIGHTS,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT, CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    CF_WARNING_FILL, CF_WARNING_FONT,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    auto_fit_columns, normalize_cell_value,
)

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ============================================================
# Load JSON data
# ============================================================
with open('/home/z/my-project/upload/rcs-canteen-backup-2026-08-03.json', 'r') as f:
    backup = json.load(f)

json_data = backup['data']

# ============================================================
# Load original Excel (to copy the original sheets)
# ============================================================
orig_wb = openpyxl.load_workbook('/home/z/my-project/upload/RCS Canteen Stock 08 26.xlsx')

# ============================================================
# Helper: build a styled data sheet
# ============================================================
def build_data_sheet(wb, sheet_name, title, headers, data_rows, col_widths=None, number_formats=None):
    """
    Build a styled data sheet following the design system.
    - title: displayed at B2
    - headers: list of column header strings
    - data_rows: list of lists (each inner list = one row of values)
    - col_widths: dict of col_index (0-based) → width
    - number_formats: dict of col_index (0-based) → format string
    """
    ws = wb.create_sheet(sheet_name)
    num_cols = len(headers)
    last_col = num_cols + 1  # B=2, so last_col = num_cols + 1

    # 1. Setup sheet basics (title at B2, margins, grid lines off)
    setup_sheet(ws, title=title, last_col=last_col)

    # 2. Write headers at row 4, starting from column B
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=4, column=col_idx + 2, value=header)
    style_header_row(ws, row_num=4, col_start=2, col_end=last_col)

    # 3. Write data rows starting at row 5
    for row_idx, row_data in enumerate(data_rows):
        excel_row = 5 + row_idx
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=excel_row, column=col_idx + 2, value=value)
            # Apply number format if specified
            if number_formats and col_idx in number_formats:
                cell.number_format = number_formats[col_idx]
        style_data_row(ws, row_num=excel_row, col_start=2, col_end=last_col, row_index=row_idx)

    # 4. Set column widths
    if col_widths:
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 2)].width = width
    else:
        auto_fit_columns(ws, min_width=10, max_width=32, header_row=4, data_start_row=5)

    # 5. Freeze panes (freeze header row)
    ws.freeze_panes = 'B5'

    return ws


# ============================================================
# Create new workbook
# ============================================================
wb = Workbook()
# Remove the default sheet
wb.remove(wb.active)

# ============================================================
# 1. Copy original sheets (Master, Purchase, Use, Meal, Stock Report)
# ============================================================
for sheet_name in ['Master', 'Purchase', 'Use', 'Meal', 'Stock Report']:
    if sheet_name in orig_wb.sheetnames:
        orig_ws = orig_wb[sheet_name]
        new_ws = wb.create_sheet(sheet_name)
        # Copy cell values, formulas, and basic formatting
        for row in orig_ws.iter_rows(min_row=1, max_row=orig_ws.max_row, max_col=orig_ws.max_column):
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.border = copy(cell.border)
                    new_cell.number_format = cell.number_format
        # Copy merged cells
        for merged_range in orig_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        # Copy column widths
        for col_letter, dim in orig_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width
        # Copy row heights
        for row_num, dim in orig_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height

# ============================================================
# 2. SUPPLIER MASTER (Missing from Excel)
# ============================================================
suppliers = json_data.get('Supplier', [])
supplier_headers = ['No', 'Supplier Name', 'Category', 'Contact Person', 'Phone', 'Email', 'Address', 'GSTIN', 'Rating', 'On-Time Rate', 'Quality Score', 'Notes', 'Last Order Date']
supplier_rows = []
for i, s in enumerate(suppliers, 1):
    supplier_rows.append([
        i,
        s.get('name', ''),
        s.get('category', ''),
        s.get('contactPerson') or '',
        s.get('phone') or '',
        s.get('email') or '',
        s.get('address') or '',
        s.get('gstin') or '',
        s.get('rating'),
        s.get('onTimeRate'),
        s.get('qualityScore'),
        s.get('notes') or '',
        s.get('lastOrderDate') or '',
    ])
build_data_sheet(
    wb, 'Supplier Master', 'RCS Canteen — Supplier Master',
    supplier_headers, supplier_rows,
    col_widths={0: 6, 1: 20, 2: 14, 3: 18, 4: 16, 5: 24, 6: 28, 7: 18, 8: 10, 9: 14, 10: 14, 11: 28, 12: 16},
    number_formats={8: '0.0', 9: '0.0%', 10: '0.0'}
)

# ============================================================
# 3. INGREDIENT MASTER (Enhanced — missing from Excel)
# ============================================================
ingredients = json_data.get('Ingredient', [])
# Build a supplier lookup for name resolution
supplier_lookup = {s['id']: s['name'] for s in suppliers}

ingredient_headers = ['No', 'Ingredient', 'Unit', 'Category', 'Current Stock', 'Min Stock', 'Last Purchase Price', 'Avg Cost', 'Supplier']
ingredient_rows = []
for i, ing in enumerate(ingredients, 1):
    supplier_name = supplier_lookup.get(ing.get('supplierId'), '') or ''
    ingredient_rows.append([
        i,
        ing.get('name', ''),
        ing.get('unit', ''),
        ing.get('category', ''),
        ing.get('currentStock', 0),
        ing.get('minStock', 0),
        ing.get('lastPurchasePrice', 0),
        ing.get('avgCost', 0),
        supplier_name,
    ])
build_data_sheet(
    wb, 'Ingredient Master', 'RCS Canteen — Ingredient Master',
    ingredient_headers, ingredient_rows,
    col_widths={0: 6, 1: 18, 2: 10, 3: 14, 4: 16, 5: 14, 6: 20, 7: 14, 8: 18},
    number_formats={4: '#,##0.0', 5: '#,##0.0', 6: '#,##0.00', 7: '#,##0.00'}
)

# ============================================================
# 4. RECIPE (Missing from Excel)
# ============================================================
recipes = json_data.get('Recipe', [])
recipe_headers = ['No', 'Recipe Name', 'Description', 'Meal Type', 'Base Servings', 'Instructions']
recipe_rows = []
for i, r in enumerate(recipes, 1):
    recipe_rows.append([
        i,
        r.get('name', ''),
        r.get('description') or '',
        r.get('mealType', ''),
        r.get('baseServings', 0),
        r.get('instructions') or '',
    ])
build_data_sheet(
    wb, 'Recipe', 'RCS Canteen — Recipe Master',
    recipe_headers, recipe_rows,
    col_widths={0: 6, 1: 22, 2: 30, 3: 12, 4: 16, 5: 36}
)

# ============================================================
# 5. RECIPE INGREDIENT (Missing from Excel)
# ============================================================
recipe_ingredients = json_data.get('RecipeIngredient', [])
# Build ingredient lookup
ingredient_lookup = {ing['id']: ing['name'] for ing in ingredients}
recipe_lookup = {r['id']: r['name'] for r in recipes}

ri_headers = ['No', 'Recipe', 'Ingredient', 'Quantity', 'Unit']
ri_rows = []
for i, ri in enumerate(recipe_ingredients, 1):
    ri_rows.append([
        i,
        recipe_lookup.get(ri.get('recipeId', ''), ''),
        ingredient_lookup.get(ri.get('ingredientId', ''), ''),
        ri.get('quantity', 0),
        ri.get('unit', ''),
    ])
build_data_sheet(
    wb, 'Recipe Ingredient', 'RCS Canteen — Recipe Ingredients',
    ri_headers, ri_rows,
    col_widths={0: 6, 1: 22, 2: 18, 3: 14, 4: 10},
    number_formats={3: '#,##0.0'}
)

# ============================================================
# 6. BUDGET (Missing from Excel)
# ============================================================
budgets = json_data.get('Budget', [])
budget_headers = ['No', 'Month', 'Food Budget', 'Operating Budget', 'Total Budget', 'Alert Threshold (%)']
budget_rows = []
for i, b in enumerate(budgets, 1):
    budget_rows.append([
        i,
        b.get('month', ''),
        b.get('foodBudget', 0),
        b.get('operatingBudget', 0),
        b.get('totalBudget', 0),
        b.get('alertThreshold', 0),
    ])
build_data_sheet(
    wb, 'Budget', 'RCS Canteen — Budget',
    budget_headers, budget_rows,
    col_widths={0: 6, 1: 14, 2: 18, 3: 20, 4: 18, 5: 22},
    number_formats={2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '0'}
)

# ============================================================
# 7. EXPENSE (Missing from Excel — template with headers)
# ============================================================
expense_headers = ['No', 'Date', 'Category', 'Description', 'Amount', 'Payment Method', 'Approved By', 'Notes']
expense_rows = []  # Empty — no expense records in JSON
build_data_sheet(
    wb, 'Expense', 'RCS Canteen — Expenses',
    expense_headers, expense_rows,
    col_widths={0: 6, 1: 14, 2: 16, 3: 28, 4: 14, 5: 16, 6: 16, 7: 28}
)

# ============================================================
# 8. DAILY MEAL SERVED (Missing from Excel — template with headers)
# ============================================================
dms_headers = ['No', 'Date', 'Meal Type', 'Count', 'Notes']
dms_rows = []  # Empty — no records in JSON
build_data_sheet(
    wb, 'Daily Meal Served', 'RCS Canteen — Daily Meals Served',
    dms_headers, dms_rows,
    col_widths={0: 6, 1: 14, 2: 14, 3: 10, 4: 28}
)

# ============================================================
# 9. STOCK MOVEMENT (Missing from Excel — template with headers)
# ============================================================
sm_headers = ['No', 'Date', 'Ingredient', 'Movement Type', 'Quantity', 'Unit', 'Reference', 'Notes']
sm_rows = []  # Empty — no records in JSON
build_data_sheet(
    wb, 'Stock Movement', 'RCS Canteen — Stock Movement Log',
    sm_headers, sm_rows,
    col_widths={0: 6, 1: 14, 2: 18, 3: 16, 4: 14, 5: 10, 6: 18, 7: 28}
)

# ============================================================
# 10. AUDIT LOG (Missing from Excel)
# ============================================================
audit_logs = json_data.get('AuditLog', [])
audit_headers = ['No', 'Date', 'User', 'Role', 'Action', 'Entity Type', 'Entity Name', 'Description']
audit_rows = []
for i, a in enumerate(audit_logs, 1):
    audit_rows.append([
        i,
        a.get('createdAt', ''),
        a.get('userName') or 'System',
        a.get('userRole') or '',
        a.get('action', ''),
        a.get('entityType', ''),
        a.get('entityName') or '',
        a.get('description', ''),
    ])
build_data_sheet(
    wb, 'Audit Log', 'RCS Canteen — Audit Log',
    audit_headers, audit_rows,
    col_widths={0: 6, 1: 20, 2: 14, 3: 12, 4: 12, 5: 14, 6: 14, 7: 36}
)

# ============================================================
# 11. USER (Missing from Excel)
# ============================================================
users = json_data.get('User', [])
user_headers = ['No', 'Name', 'Email', 'Role']
user_rows = []
for i, u in enumerate(users, 1):
    user_rows.append([
        i,
        u.get('name', ''),
        u.get('email', ''),
        u.get('role', ''),
    ])
build_data_sheet(
    wb, 'User', 'RCS Canteen — Users',
    user_headers, user_rows,
    col_widths={0: 6, 1: 22, 2: 28, 3: 12}
)

# ============================================================
# Set workbook properties
# ============================================================
wb.properties.creator = "Z.ai"
wb.properties.title = "RCS Canteen Stock — Enhanced"
wb.properties.subject = "RCS Canteen Management — Complete Data from JSON Backup"

# ============================================================
# Save
# ============================================================
output_path = '/home/z/my-project/download/RCS_Canteen_Stock_Enhanced.xlsx'
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Enhanced Excel saved to: {output_path}")

# ============================================================
# Print summary
# ============================================================
print("\n=== SHEET SUMMARY ===")
for ws in wb.worksheets:
    print(f"  {ws.title}: {ws.max_row} rows x {ws.max_column} cols")

print("\n=== WHAT WAS MISSING IN ORIGINAL EXCEL (NOW ADDED) ===")
missing_items = [
    ("Supplier Master", "3 suppliers with contact, GSTIN, category, rating, quality scores"),
    ("Ingredient Master", "24 ingredients with category, current stock, min stock, last price, avg cost, supplier mapping"),
    ("Recipe", "2 recipes (Chicken Curry, Dal Rice) with meal type, base servings, description"),
    ("Recipe Ingredient", "14 recipe-ingredient mappings with quantity per recipe"),
    ("Budget", "Monthly budget (5,00,000 food / 7,50,000 operating) with 80% alert threshold"),
    ("Expense", "Empty template — ready for expense tracking (category, amount, payment method, approval)"),
    ("Daily Meal Served", "Empty template — ready for daily meal count tracking"),
    ("Stock Movement", "Empty template — ready for stock movement audit trail"),
    ("Audit Log", "2 system audit entries (DB seed, budget creation)"),
    ("User", "1 admin user account"),
]
for name, desc in missing_items:
    print(f"  + {name}: {desc}")
